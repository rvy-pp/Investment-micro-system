"""L0 adapter — Kotak's Daily Cement Pack into `prices`.

    python packages/adapters/cement_pack.py --file data/staging/cement_pack_YYYY-MM-DD.tsv --probe
    python packages/adapters/cement_pack.py --file data/staging/cement_pack_YYYY-MM-DD.tsv --load

WHAT IT SUPPLIES THAT NOTHING ELSE DOES. Cement's OUTPUT price. Every cost line
the sector needs — pet coke, seaborne and Indonesian thermal coal, Brent, USDINR
— is already in the Daily METALS Pack and has been captured daily since 2010.
The revenue leg had no feed at all, which is why `engine.SECTORS['cement']`
carried `peer_groups: []` and a comment saying wiring this pack is step one.

THE CONNECTOR ROUTE WORKS HERE AND FAILS FOR THE METALS PACK. That is not a
contradiction and it is worth stating, because `outlook_pack.py` exists solely
to route around the M365 connector and a reader who knows that will assume the
same is needed here.

  metals pack   TALL — one row per date, ~4,760 rows back to 2010. The connector
                caps an extraction at ~200k characters and keeps the OLDEST
                rows, so it returns 2010-2013 and never today. Hence MAPI.
  cement pack   WIDE — dates run ACROSS the columns, so the whole workbook is
                336 lines and ~70k characters. The connector returns ALL FOUR
                sheets complete, current to the same morning.

Verified 2026-08-27: 70,034 characters, well inside the cap, ending on a real
trailer row rather than a truncation footer. So this one is an ordinary
connector read and needs no Outlook automation. If the pack ever grows past the
cap the symptom will be a SHORT span rather than an error — `--probe` prints the
span for exactly that reason.

THE FILENAME CARRIES A STRAY SPACE. `Daily Cement Pack, August 27 , 2026.xlsx`.
`outlook_pack.py` records that this is NOT a stable discriminator — on
2026-08-18 both packs had it. Match the `Daily Cement Pack` PREFIX.

--- two things in the sheet that would each have loaded a wrong number ---

1. THE DATE CONVENTION CHANGES MID-SERIES, at 2019-10-01 -> 2019-11-30. The
   first 67 columns label a month by its FIRST day and the remaining 82 by its
   LAST. Loaded verbatim that is not a gap and not a crash — every month is
   still present exactly once — but the store would hold two different meanings
   for the same field and any later month-arithmetic would be off by one at the
   seam. Every column is therefore normalised to (year, month) and stamped at
   MONTH END.

   Month end, not month start, and the reason is the look-ahead rule in
   CLAUDE.md: a dated value carries the date the market COULD HAVE KNOWN it. An
   average of August's prices is knowable on 31 August, not on the 1st. Note
   `iron_ore` (FRED/IMF monthly) is dated on the 1st and has the opposite
   convention; that is a pre-existing series and is not touched here, but it
   means the store now contains both and neither is universally right.

1b. AND THE CURRENT MONTH IS NOT A COMPLETED MONTH. Caught by running the load
   and then looking at MAX(date), which is the only reason it was caught at all.
   The pack prints an August column on 27 August — a month-TO-DATE average — and
   stamping it 2026-08-31 wrote six rows FOUR DAYS IN THE FUTURE. `bridge.py`
   takes its default `as_of` from `MAX(date) FROM prices`, so that one stamp
   moved every pillar's as_of forward four days, silently, with nothing raising:
   the exact shape of `docs/SILENT_BUGS.md`. `westmetall.py` already refuses a
   future date for the same reason and this had no equivalent.

   THE RULE NOW: stamp at `min(month_end, as_of)`. A completed month lands on
   its month end and never moves again; the in-progress month lands on the
   capture date and walks forward each morning until the month closes. `as_of`
   comes from the staging FILENAME (`cement_pack_YYYY-MM-DD`), not from
   `today()`, so re-loading an old capture reproduces what that morning knew
   rather than re-dating it to now. A future `as_of` is refused outright.

   The consequence, stated because it is a real cost: loading twice in one month
   would otherwise leave a TRAIL of partial-August rows, so `--load` deletes the
   in-progress month's rows for these six ids before writing. That discards the
   record of what August's MTD average read on the 26th. It is not lost — the
   dated staging .tsv holds it, which is what staging captures are for.

1c. AND "DAILY" IS THE MAIL, NOT THE PRICES. Measured, not assumed: the
   2026-08-27 and 2026-08-28 captures were diffed and the Monthly Prices sheet
   is IDENTICAL to the fourth decimal — all six regions, all 149 months, the
   in-progress August column included. The pack is mailed every morning; the
   price series inside it is monthly.

   So the walking stamp from 1b must NOT walk on an unchanged value. Moving the
   August row from 08-27 to 08-28 because a mail arrived would put a fresh
   print date on a number nobody re-observed — `series.py` rule 2 in as many
   words ("a restatement is not a new datapoint"), and CLAUDE.md invariant 3
   ("silence changes nothing"). `--load` therefore leaves the stored row alone
   when the value matches, and only a CHANGED value re-stamps the month.

   The update cadence of the current-month column is still UNKNOWN — two
   adjacent captures cannot distinguish weekly from month-end. It does not need
   to be known for the load to be correct, because the rule keys on the value
   rather than on a predicted schedule.

   BOTH BRANCHES ARE TESTED, per the GLOB lesson that a guard needs an
   acceptance test and not only a rejection test. Loading the 08-28 capture over
   the 08-27 one KEEPS the row at 08-27 (149 rows, no drift); loading the same
   file with +Rs2/bag injected into the August column CLEARS and re-stamps it at
   08-28 (149 rows, +Rs40/t). The second branch is the one that matters more —
   if it were broken the current month would freeze at its first capture date
   and go on reporting a stale price with nothing raising.

2. THE SHEET REPEATS THE SIX REGION LABELS THREE TIMES — levels, then
   `Change per bag (Rs, mom)`, then `(Rs, yoy)`. A naive label match takes the
   LAST block and loads month-on-month DELTAS as though they were prices, which
   would be numbers around +/-10 in a field the bridge reads as a level. Two
   independent guards: only the first block after the header is taken, and every
   value must fall in PLAUSIBLE_BAG. A delta block fails the range check even if
   the block logic were wrong.

--- units: stored in Rs/TONNE, and this is load-bearing ---

The sheet prints Rs per 50 kg bag. `bridge.py` multiplies an ABSOLUTE price
delta by a tonnage:

    d_cost = basis_volume * intensity * market_pct * d_price

so a Rs/bag delta against a tonnage basis understates the revenue leg by
EXACTLY 20x, silently, with `coverage_ok` still true. That is the failure shape
`docs/SILENT_BUGS.md` catalogues — entry 3 is an unregistered unit that dropped
an FX leg 95x. Rs/t also matches the convention the steel group already uses
(`hrc_india_inr`, `rebar_india_primary_inr` are both Rs/t), so a cement line and
a steel line can sit in one spec without a per-line unit.

The staging .tsv keeps the source's own Rs/bag numbers, so the conversion is
auditable rather than lost. TO CONVERT BACK, DIVIDE BY 20.

--- OCTOBER 2025 IS A REAL MOVE. DO NOT CLEAN IT. ---

All-India -9.37% in one month, every region -8.0% to -11.4%. It is the largest
monthly move in 149 months by more than 2x (next worst -4.43%) and October's own
seasonal mean is -0.16%, so every statistical property of a rebasing is present.
**PM ruling 2026-08-28: it is genuine.** No exclusion window, no allow-list, no
cleaning. A P1 window spanning Sep->Oct 2025 SHOULD book it. This note exists
because the next person to run an outlier scan will find it and reach for the
same wrong conclusion I did.

--- THE PACK LANDS ~15 DAYS LATE ---

PM figure, 2026-08-28. Cement hikes go in at the start of a month, so a
fortnight's lag means this series CONFIRMS a move rather than carrying it while
it is tradeable. That is not a defect in the pack and nothing here tries to
correct for it — it is why `adapters/indiamart_cement.py` exists alongside it as
a same-day move detector. Keep the two straight: this file is the priced series
and the reference; that one is a watch that never reaches a score.

--- the level disagrees with Nomura's and that is NOT resolved ---

Kotak's all-India average reads 353.0 in Jul-2026 and 351.4 in Aug-2026. The
broker digests carry Nomura channel checks at ~Rs326/bag in Jun and ~Rs321/bag
in Jul — roughly Rs30/bag, ~9%, below. Nomura says explicitly "trade prices";
this sheet says only "cement prices", so the likely difference is trade versus a
trade/non-trade blend, or a different dealer panel.

It is NOT aliased to either. Per invariant 6 a proxy is never given the name of
the thing it proxies, so these ids say `cement_price_<region>_inr` and claim no
basis. The bridge reads DELTAS, so a constant level offset does not reach a
score; a divergence in the CHANGES would, and would mean the two panels are
measuring different markets. Worth checking before any spec cites a level.

--- what this deliberately does NOT load ---

The pack has four sheets. Only `Monthly Prices` is loaded.

  Volumes     DIPP all-India monthly cement production, '000 t, 268 points back
              to 2004-04. It is a DEMAND indicator, not a price, and putting a
              tonnage into `prices.close` would make `_series_in_store()` claim
              it as a priceable series the bridge could shock. It has a home in
              this system — Flows, or a P2 volume input — and no schema for it
              yet. Unlike the metals pack there is no urgency: the connector
              returns the FULL history every morning, so nothing is lost by
              waiting. See CLAUDE.md's "a series you are not capturing today is
              history you cannot recover" — that argument does not apply here.
  Valuation   Kotak's comparative table: mcap, CMP, fair value, rating, EPS,
              P/E, P/B, EV/ton of capacity, EV/EBITDA, RoE, FY26-29E. These are
              P3 inputs and they matter, because CLAUDE.md records that Wind
              returns EMPTY for every Indian fundamental field. Not loaded here
              because `prices` is the wrong table for a valuation multiple.
  Stock       1m/3m/6m/1y/YTD absolute and relative performance for nine names,
              including Grasim, which is not in the vault's cement roster.
"""

from __future__ import annotations

import argparse
import calendar
import csv
import datetime as dt
import pathlib
import sqlite3
import sys

REPO = pathlib.Path(__file__).resolve().parent.parent.parent
sys.path.insert(0, str(REPO / "packages" / "core"))

import prices_io  # noqa: E402

DB = REPO / "data" / "ims.db"
SHEET = "Monthly Prices"

# Excel's serial-date origin. The 30th, not the 31st: the 1900 leap-year bug.
EPOCH = dt.date(1899, 12, 30)

# Rs per 50 kg bag -> Rs per tonne. See the units section of the docstring; this
# constant is the whole of the 20x exposure and is deliberately not inlined.
BAG_KG = 50
PER_TONNE = 1000 / BAG_KG

# Sheet row label -> entity id. Labels are matched EXACTLY after stripping.
ROWS = {
    "North":             ("cement_price_north_inr",   "North India"),
    "Central":           ("cement_price_central_inr", "Central India"),
    "East":              ("cement_price_east_inr",    "East India"),
    "West":              ("cement_price_west_inr",    "West India"),
    "South":             ("cement_price_south_inr",   "South India"),
    "All India average": ("cement_price_india_inr",   "All-India average"),
}

# Rs/BAG, checked BEFORE conversion. The series has run 240-360 over twelve
# years; this is wide enough to admit a real move and narrow enough that a
# mom-delta block (values around +/-15) or a yoy block cannot pass.
PLAUSIBLE_BAG = (100.0, 1000.0)


def _stamp(d: dt.date, as_of: dt.date) -> str:
    """Date one month column: month end, or the capture date if still running.

    See point 1b of the module docstring. The `min` is the whole guard against
    writing a future row, and it must not be simplified back to a plain month
    end — a month-to-date average dated at month end is both a look-ahead and a
    future date, and the future date silently moves bridge.py's default as_of.
    """
    end = d.replace(day=calendar.monthrange(d.year, d.month)[1])
    return min(end, as_of).isoformat()


def as_of_from_name(path: pathlib.Path, override: str | None = None) -> dt.date:
    """Capture date: --as-of, else the staging filename, else today.

    From the FILENAME rather than today() so re-running an old capture
    reproduces what that morning knew. A capture with no date in its name falls
    back to today with a warning, because guessing silently is how a look-ahead
    gets in.
    """
    if override:
        return dt.date.fromisoformat(override)
    m = [p for p in path.stem.split("_") if len(p) == 10 and p[4] == "-"]
    if m:
        return dt.date.fromisoformat(m[-1])
    print(f"WARN: no YYYY-MM-DD in {path.name}; dating the capture as today. "
          f"Pass --as-of to be exact.", file=sys.stderr)
    return dt.date.today()


def read_tsv(path: pathlib.Path, as_of: dt.date) -> dict[str, dict[str, float]]:
    """Parse the connector's text extraction. Returns {entity_id: {iso: Rs/t}}.

    The extraction separates sheets with `=== Sheet: <name> ===` lines, and the
    Monthly Prices sheet is WIDE: a header row of Excel serials, then one row
    per region. Everything is keyed off finding that header, so a layout change
    fails loudly here rather than producing a short series.
    """
    with open(path, encoding="utf-8") as f:
        rows = list(csv.reader(f, delimiter="\t"))

    marker = "=== Sheet: " + SHEET + " ==="
    start = next((i for i, r in enumerate(rows)
                  if r and r[0].strip() == marker), None)
    if start is None:
        present = [r[0] for r in rows if r and r[0].startswith("=== Sheet")]
        raise ValueError("no " + marker + " in " + path.name
                         + "; sheets present: " + repr(present))
    end = next((i for i in range(start + 1, len(rows))
                if rows[i] and rows[i][0].startswith("=== Sheet:")), len(rows))
    block = rows[start:end]

    # The header is the first row with a blank label and a long run of serials.
    hdr_i, dates = None, []
    for i, r in enumerate(block):
        if not r or r[0].strip():
            continue
        ds = []
        for cell in r[1:]:
            cell = cell.strip()
            if not cell:
                ds.append(None)
                continue
            try:
                serial = int(float(cell))
            except ValueError:
                ds.append(None)
                continue
            # ~1954..2064. Rejects a stray count or percentage sitting in the
            # header band, which is how a wide sheet usually goes wrong.
            ds.append(EPOCH + dt.timedelta(days=serial)
                      if 20000 < serial < 60000 else None)
        if sum(x is not None for x in ds) >= 24:
            hdr_i, dates = i, ds
            break
    if hdr_i is None:
        raise ValueError("no date header found in sheet " + SHEET
                         + " of " + path.name)

    out: dict[str, dict[str, float]] = {eid: {} for eid, _ in ROWS.values()}
    seen: set[str] = set()
    for r in block[hdr_i + 1:]:
        if not r:
            continue
        label = r[0].strip()
        # FIRST block only — the same six labels reappear under the mom and yoy
        # change headers further down. See guard 2 in the docstring.
        if label not in ROWS or label in seen:
            continue
        seen.add(label)
        eid, _name = ROWS[label]
        for j, d in enumerate(dates, start=1):
            if d is None or j >= len(r):
                continue
            raw = r[j].strip()
            if not raw or raw.startswith("#"):        # '#N/A' is a gap
                continue
            try:
                v = float(raw)
            except ValueError:
                continue
            if not PLAUSIBLE_BAG[0] <= v <= PLAUSIBLE_BAG[1]:
                raise ValueError(
                    eid + " " + d.isoformat() + ": " + format(v, "g")
                    + " Rs/bag is outside " + repr(PLAUSIBLE_BAG)
                    + " — this is what a delta block read as a level looks"
                      " like. Check the sheet layout before widening it.")
            out[eid][_stamp(d, as_of)] = v * PER_TONNE
    missing = [lbl for lbl in ROWS if lbl not in seen]
    if missing:
        raise ValueError("labels not found in sheet " + SHEET + ": "
                         + repr(missing))
    return out


def read_xlsx(path: pathlib.Path, as_of: dt.date) -> dict[str, dict[str, float]]:
    """Same sheet out of a real workbook, for a hand-dropped or MAPI-saved file.

    Present so the connector is not a single point of failure — if the pack ever
    outgrows the ~200k cap, `outlook_pack.py` can save the bytes and this reads
    them with no other change. Deliberately mirrors read_tsv's guards rather
    than sharing them, because openpyxl hands back real dates and floats and the
    string-parsing half does not apply.
    """
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if SHEET not in wb.sheetnames:
        raise ValueError("no sheet " + SHEET + " in " + path.name
                         + "; sheets: " + repr(wb.sheetnames))
    grid = [list(r) for r in wb[SHEET].iter_rows(values_only=True)]

    hdr_i, dates = None, []
    for i, r in enumerate(grid):
        if r and r[0] not in (None, ""):
            continue
        ds = [c.date() if isinstance(c, dt.datetime)
              else (c if isinstance(c, dt.date) else None) for c in r[1:]]
        if sum(x is not None for x in ds) >= 24:
            hdr_i, dates = i, ds
            break
    if hdr_i is None:
        raise ValueError("no date header found in sheet " + SHEET
                         + " of " + path.name)

    out: dict[str, dict[str, float]] = {eid: {} for eid, _ in ROWS.values()}
    seen: set[str] = set()
    for r in grid[hdr_i + 1:]:
        label = str(r[0]).strip() if r and r[0] is not None else ""
        if label not in ROWS or label in seen:
            continue
        seen.add(label)
        eid, _name = ROWS[label]
        for j, d in enumerate(dates, start=1):
            v = r[j] if d is not None and j < len(r) else None
            if not isinstance(v, (int, float)) or isinstance(v, bool):
                continue
            if not PLAUSIBLE_BAG[0] <= float(v) <= PLAUSIBLE_BAG[1]:
                raise ValueError(
                    eid + " " + d.isoformat() + ": " + format(v, "g")
                    + " Rs/bag is outside " + repr(PLAUSIBLE_BAG)
                    + " — see read_tsv for why this guard exists.")
            out[eid][_stamp(d, as_of)] = float(v) * PER_TONNE
    missing = [lbl for lbl in ROWS if lbl not in seen]
    if missing:
        raise ValueError("labels not found in sheet " + SHEET + ": "
                         + repr(missing))
    return out


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--file", required=True,
                    help=".tsv/.txt from the M365 connector, or an .xlsx")
    ap.add_argument("--as-of", default=None,
                    help="capture date YYYY-MM-DD; default: from the filename")
    ap.add_argument("--probe", action="store_true")
    ap.add_argument("--load", action="store_true")
    a = ap.parse_args()

    f = pathlib.Path(a.file)
    as_of = as_of_from_name(f, a.as_of)
    if as_of > dt.date.today():
        print(f"REFUSED: capture date {as_of} is in the future. Dating a "
              f"price forward is a look-ahead, and it also moves bridge.py's "
              f"default as_of, which nothing else would report.",
              file=sys.stderr)
        return 1
    print(f"capture date {as_of} — the in-progress month "
          f"({as_of:%Y-%m}) is stamped here, completed months at month end\n")
    series = (read_tsv(f, as_of) if f.suffix.lower() in (".tsv", ".txt")
              else read_xlsx(f, as_of))

    if a.probe or not a.load:
        print("entity                        rows  span                        "
              " latest Rs/t   Rs/bag")
        for _lbl, (eid, name) in ROWS.items():
            ks = sorted(series[eid])
            span = (ks[0] + " .. " + ks[-1]) if ks else "— none in this file —"
            last = series[eid][ks[-1]] if ks else float("nan")
            print(f"{eid:28}{len(ks):>6}  {span:<28}{last:>12,.0f}"
                  f"{last / PER_TONNE:>9.1f}   {name}")
        if not a.load:
            print("\nprobe only — pass --load to write")
            return 0

    conn = sqlite3.connect(DB)
    conn.execute("PRAGMA foreign_keys = ON")
    n = 0
    cleared = 0
    carried = 0
    for _lbl, (eid, name) in ROWS.items():
        # is_tradeable=1 because the schema demands a parent_id otherwise and a
        # bare commodity has none; peer_group stays NULL so it is never scored.
        label = ("Cement price, " + name + ", Rs/t "
                 "(Kotak pack; source prints Rs/50kg bag)")
        conn.execute(
            "INSERT OR IGNORE INTO entities (id,kind,name,is_tradeable,active) "
            "VALUES (?,'commodity',?,1,1)", (eid, label))
        # The IN-PROGRESS month, and the rule is NOT "restamp it every morning".
        # See docstring 1c: the pack is mailed daily but the price sheet is
        # monthly, so on most mornings the current month's number is the SAME
        # number. Re-dating it to today would manufacture a fresh print out of
        # an unchanged value — which series.py rule 2 forbids in as many words
        # ("a restatement is not a new datapoint") and CLAUDE.md invariant 3
        # restates as "silence changes nothing".
        #
        # So: leave the existing row exactly where it is when the value has not
        # moved. Only a CHANGED value clears the month and re-stamps it at the
        # capture date. Scoped to this month and these six ids; completed months
        # are never touched, so a re-load is idempotent everywhere else.
        month_start = as_of.replace(day=1).isoformat()
        incoming = {d: c for d, c in series[eid].items() if d >= month_start}
        held = conn.execute(
            "SELECT date, close FROM prices WHERE entity_id=? AND date>=? "
            "AND date<=?", (eid, month_start, as_of.isoformat())).fetchall()
        unchanged = (len(held) == 1 and len(incoming) == 1
                     and abs(held[0][1] - next(iter(incoming.values()))) < 1e-9)
        if unchanged:
            held_date = held[0][0]
            series[eid] = {d: c for d, c in series[eid].items()
                           if d < month_start}
            series[eid][held_date] = held[0][1]
            carried += 1
        elif held:
            cleared += conn.execute(
                "DELETE FROM prices WHERE entity_id=? AND date>=? AND date<=?",
                (eid, month_start, as_of.isoformat())).rowcount
        rows = [(eid, d, c) for d, c in sorted(series[eid].items())]
        assert all(d <= as_of.isoformat() for _e, d, _c in rows), \
            f"{eid}: a row is dated after the capture date — _stamp() is broken"
        res = prices_io.upsert(conn, rows, "cement_pack")
        n += res["wrote"]
        if res.get("refused"):
            print(f"  {eid}: {res['refused']:,} REFUSED (higher-ranked source)")
    conn.commit()
    conn.close()
    if carried:
        print(f"{carried} series: {as_of:%Y-%m} unchanged since its stored date "
              f"— kept, NOT re-stamped to {as_of}")
    if cleared:
        print(f"cleared {cleared:,} in-progress rows for {as_of:%Y-%m} (value moved)")
    print(f"loaded {n:,} price rows into {DB}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
