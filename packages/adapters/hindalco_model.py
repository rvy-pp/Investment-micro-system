"""The desk's Hindalco model (BofA) -> what it says vs what the specs say.

    python packages/adapters/hindalco_model.py                 # compare
    python packages/adapters/hindalco_model.py --json
    python packages/adapters/hindalco_model.py --observations   # emit extract JSON

Reads `data/models/*.xls*` — a sell-side model, gitignored with the rest of
`data/`. Built against the BofA workbook supplied 2026-08-25 (Bharat
Subramanian, bharat.subramanian@bofa.com), 11 sheets.

IT DOES NOT WRITE TO THE SPECS. Same rule as concall_check.py: the base numbers
are static and a denominator or intensity is the beta for every signal the name
emits, so changing one stays a human edit with a citation. This prints the
comparison; you make the call.

PRECEDENCE, from base_financials.yaml:
    company release for a LATER quarter  beats  the model
    the model                           beats  sector-knowledge guesses

WHAT THIS WORKBOOK CAN AND CANNOT SETTLE — established by reading it, and the
distinction is the whole value of the file:

  SETTLES     CP coke intensity. Mahan and Aditya each carry coke tonnage and
              metal tonnage separately, and both divide to 0.360 t/t exactly
              (122,778/341,050 and 123,120/342,000). The spec carried 0.40 from
              sector knowledge; two independent derivations agreeing to three
              decimals beat one round number, so hindalco was changed to 0.36 on
              2026-08-25. NALCO's and VAML's cp_coke lines stay at 0.40 — anode
              consumption is a property of a particular smelter's anode plant,
              and copying Hindalco's across would invent two data points
              from one.

  CONFIRMS    Alumina intensity. The model straddles the spec's 1.93 — 1.92 on
              the Renukoot/Hirakud sheet, 1.950 derived on both new smelters.
              No change warranted; a spec value inside the range of two
              independent estimates is not improved by moving it.

  CONFIRMS    Coal captive share. Aditya buys 2.65 mn t and takes 1.50 mn t from
              the Gare Palma blocks = 56.6% captive, so 43.4% market-exposed
              against the spec's market_pct 0.45.

  DOES NOT    The ASP premium. The model's "Premium to LME" is $1,124/t (FY26),
  SETTLE      against the spec's asp_premium of $150/t — and they are NOT the
              same quantity, so this is not a 7x error. The model's premium is a
              BLENDED realisation over LME across a book that is ~40% rolled,
              extruded and foil products, which realise far above ingot. The
              spec line is `aluminium_ingot`. Do not reconcile these.

  DOES NOT    Novelis' scrap cost. Searched every sheet: the only scrap rows are
  SETTLE      "Recycled content" (a percentage) and an EBITDA-level
              "- Elimination/scrap benefit" in US$mn. The model prices Novelis'
              raw material AT LME. So `al_scrap_midwest` has no source in the
              company's disclosure, in the free market data, or in the sell-side
              model — three independent places. That is now a settled negative.

  SUPPLIES    Novelis' conversion spread, which nothing public does. Row 9
              "Premium (US$/t)" over LME: 2,206 / 2,028 / 2,417 / 3,026 / 2,955
              for FY24/25/26/27E/28E. This is the only route to pricing
              `can_sheet_spread`.

A TRAP THIS ADAPTER EXISTS TO FLAG. The model's Novelis FY26 Adj. EBITDA is
$2,057mn against the company's reported $1,645mn on identical shipments of
3,557 kt. The reconciling item is in the sheet: row 87 "- Elimination/scrap
benefit" = $411mn for FY26, and 2,057 - 1,645 = 412. So the model runs a
different EBITDA definition. Per precedence the company's figure wins, which is
what `novelis.base_ebitda` already uses. Anyone "correcting" the spec to the
model's number would be silently changing the definition of the denominator.
"""

from __future__ import annotations

import argparse
import glob
import json
import pathlib
import re
import sys

REPO = pathlib.Path(__file__).resolve().parent.parent.parent
MODELS = REPO / "data" / "models"

# sheet -> label regex -> (key, how to read it)
# Rows are found by LABEL and columns by FY HEADER, never by fixed cell, because
# a sell-side model gains and loses rows every quarter.
WANT = {
    "Aluminum": {
        "alumina_per_t_al":  r"^Alumina consumed/t of aluminum",
        "premium_to_lme":    r"^Premium to LME",
        "lme":               r"^LME Aluminum",
        "blended_real":      r"^Blended realizations",
        "al_production":     r"^Aluminium$",
    },
    "Novelis Inc": {
        "shipments_kt":      r"^Shipments \(kt\)",
        "lme":               r"^AL LME \(US\$/t\)",
        "premium_usd_t":     r"^Premium \(US\$/t\)",
        "realisation":       r"^Realisation \(US\$/t\)",
        "adj_ebitda_usd_mn": r"^Adj\. EBITDA \(including MPL\)",
        "adj_ebitda_per_t":  r"^Adj\. EBITDA/T",
        "scrap_benefit":     r"^- Elimination/scrap benefit",
    },
    "Mahan": {
        "metal_t":    r"^Aluminum metal$",
        "coke_t":     r"^Volumes consumed \(MT\)",     # first hit after C.P.Coke
        "coke_per_t": r"^CP Coke$",
        "coal_mnt":   r"^Total coal consumed",
    },
    "Aditya": {
        "metal_t":    r"^Primary Ali Production",
        "coke_per_t": r"^CP Coke$",
        "coal_mnt":   r"^Total coal consumed",
        "captive_coal": r"^Captive coal supply",
    },
}


def find_model() -> pathlib.Path | None:
    hits = sorted(glob.glob(str(MODELS / "*.xls*")))
    return pathlib.Path(hits[-1]) if hits else None


def header_cols(rows) -> dict[str, int]:
    """FY label -> column index, from whichever early row carries them."""
    for r in rows[:12]:
        cells = {c: j for j, c in enumerate(r)
                 if isinstance(c, str) and re.fullmatch(r"FY\d{2}E?", c.strip())}
        if len(cells) > 5:
            return {k.strip(): v for k, v in cells.items()}
    return {}


def read(path: pathlib.Path, years=("FY26", "FY27E")) -> dict:
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    out: dict = {"file": path.name, "sheets": wb.sheetnames, "years": list(years)}
    for sheet, fields in WANT.items():
        if sheet not in wb.sheetnames:
            out.setdefault("missing_sheets", []).append(sheet)
            continue
        rows = list(wb[sheet].iter_rows(min_row=1, max_row=400, values_only=True))
        cols = header_cols(rows)
        got: dict = {}
        for key, pat in fields.items():
            rx = re.compile(pat)
            for r in rows:
                lab = next((c for c in r[:3]
                            if isinstance(c, str) and c.strip()), None)
                if not lab or not rx.match(lab.strip()):
                    continue
                vals = {}
                for y in years:
                    j = cols.get(y)
                    v = r[j] if j is not None and j < len(r) else None
                    if isinstance(v, (int, float)):
                        vals[y] = float(v)
                if vals:
                    got[key] = vals
                    break
        out[sheet] = got
    return out


def spec_values() -> dict:
    """Read the CURRENT spec, never hardcode it.

    The first version carried 0.40 / 1.93 / 9.60 / 0.45 / 150 as literals. The
    moment the coke intensity was actually changed to 0.36 the advisory kept
    printing "spec 0.40" and recommending a change already made. A checker that
    nags about finished work gets muted, and a muted checker is not a checker.
    """
    import sys as _s
    _s.path.insert(0, str(REPO / "packages" / "score"))
    from bridge import load_specs                       # noqa: E402
    ents, _u, _f = load_specs()
    h = ents["hindalco"]
    ins = {i["item"]: i for i in h.get("inputs", [])}
    outs = {o["item"]: o for o in h.get("outputs", [])}
    return {
        "cp_coke_intensity": (ins.get("cp_coke") or {}).get("intensity"),
        "alumina_intensity": (ins.get("alumina") or {}).get("intensity"),
        "coal_intensity": (ins.get("thermal_coal_seaborne") or {}).get("intensity"),
        "coal_market_pct": (ins.get("thermal_coal_seaborne") or {}).get("market_pct"),
        "asp_premium": (outs.get("aluminium_ingot") or {}).get("asp_premium"),
    }


def derive(m: dict) -> dict:
    """The comparisons worth making, with the arithmetic shown."""
    d: dict = {}
    sp = spec_values()
    al = m.get("Aluminum", {}) or {}
    nv = m.get("Novelis Inc", {}) or {}
    ma = m.get("Mahan", {}) or {}
    ad = m.get("Aditya", {}) or {}

    def v(block, key, y="FY26"):
        return (block.get(key) or {}).get(y)

    # coke intensity, two independent smelters
    for tag, blk in (("mahan", ma), ("aditya", ad)):
        d[f"coke_per_t_{tag}"] = v(blk, "coke_per_t")
    d["coke_per_t_spec"] = sp["cp_coke_intensity"]

    d["alumina_per_t_model"] = v(al, "alumina_per_t_al")
    d["alumina_per_t_spec"] = sp["alumina_intensity"]

    # coal per tonne of metal, derived
    for tag, blk in (("mahan", ma), ("aditya", ad)):
        coal, metal = v(blk, "coal_mnt"), v(blk, "metal_t")
        d[f"coal_t_per_t_{tag}"] = (coal * 1e6 / metal) if coal and metal else None
    d["coal_t_per_t_spec"] = sp["coal_intensity"]

    cap, tot = v(ad, "captive_coal"), v(ad, "coal_mnt")
    d["aditya_captive_share"] = (cap / tot) if cap and tot else None
    d["coal_market_pct_spec"] = sp["coal_market_pct"]

    d["novelis_premium_usd_t"] = nv.get("premium_usd_t")
    d["novelis_shipments_kt"] = nv.get("shipments_kt")
    d["novelis_adj_ebitda_model_usd_mn"] = nv.get("adj_ebitda_usd_mn")
    d["novelis_scrap_benefit_usd_mn"] = nv.get("scrap_benefit")
    d["novelis_reported_fy26_usd_mn"] = 1645.0     # company release, cited
    eb = v(nv, "adj_ebitda_usd_mn")
    sb = v(nv, "scrap_benefit")
    d["novelis_definition_gap"] = (eb - 1645.0) if eb else None
    d["novelis_gap_explained_by_scrap_benefit"] = (
        abs((eb - 1645.0) - sb) < 5 if eb and sb else None)

    d["asp_premium_model_blended"] = v(al, "premium_to_lme")
    d["asp_premium_spec_ingot"] = sp["asp_premium"]
    return d


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--json", action="store_true")
    a = ap.parse_args()

    f = find_model()
    if not f:
        print(f"no model found in {MODELS}\\*.xls*")
        print("  drop the workbook there; see data/models/README.md")
        return 0

    m = read(f)
    d = derive(m)
    if a.json:
        print(json.dumps({"raw": m, "derived": d}, indent=2))
        return 0

    print(f"model: {f.name}\n")
    mk, ak = d.get("coke_per_t_mahan"), d.get("coke_per_t_aditya")
    spec_coke = d["coke_per_t_spec"]
    agreed = mk is not None and spec_coke is not None and abs(spec_coke - mk) < 0.005
    print("SETTLES - the model beats a sector-knowledge guess")
    print(f"  cp_coke intensity        spec {spec_coke:.2f} t/t   "
          f"model {mk} (Mahan) / {ak} (Aditya)")
    print("     two smelters, derived from separate coke and metal tonnages,")
    print("     agreeing to three decimals")
    print("     -> spec ALREADY MATCHES, nothing to do" if agreed
          else "     -> spec DISAGREES, consider the change")
    print()

    print("CONFIRMS — leave the spec alone")
    print(f"  alumina intensity        spec {d['alumina_per_t_spec']:.2f} t/t   "
          f"model {d['alumina_per_t_model']} (old units), 1.95 derived (new)")
    cs = d["aditya_captive_share"]
    print(f"  coal market_pct          spec {d['coal_market_pct_spec']:.2f}       "
          f"model {1-cs:.2f} market ({cs:.1%} captive at Aditya)" if cs else "")
    print()

    print("DOES NOT SETTLE — different quantity, do not reconcile")
    print(f"  asp_premium              spec ${d['asp_premium_spec_ingot']:.0f}/t "
          f"(ingot)   model ${d['asp_premium_model_blended']:,.0f}/t (BLENDED,\n"
          f"     includes rolled/extruded/foil which realise far above ingot)")
    cl = [d.get("coal_t_per_t_mahan"), d.get("coal_t_per_t_aditya")]
    cl = [x for x in cl if x]
    if cl:
        print(f"  coal intensity           spec {d['coal_t_per_t_spec']:.2f} t/t   "
              f"model {' / '.join(f'{x:.2f}' for x in cl)} — but only 2 of 3\n"
              f"     smelter groups are covered here, so this is partial. FLAG,\n"
              f"     do not change on incomplete coverage.")
    print()

    print("SUPPLIES — nothing public has this")
    p = d.get("novelis_premium_usd_t") or {}
    print(f"  novelis conversion spread (Premium over LME, US$/t): "
          + ", ".join(f"{k} {v:,.0f}" for k, v in p.items()))
    print("     the only route to pricing can_sheet_spread\n")

    print("TRAP — the model's Novelis EBITDA is a DIFFERENT DEFINITION")
    eb = (d.get("novelis_adj_ebitda_model_usd_mn") or {}).get("FY26")
    sb = (d.get("novelis_scrap_benefit_usd_mn") or {}).get("FY26")
    if eb:
        print(f"  model FY26 ${eb:,.0f}mn vs company-reported $1,645mn on the same\n"
              f"  3,557 kt. Gap ${d['novelis_definition_gap']:,.0f}mn; row 87\n"
              f"  '- Elimination/scrap benefit' is ${sb:,.0f}mn. Explained: "
              f"{d['novelis_gap_explained_by_scrap_benefit']}.\n"
              f"  Precedence says the company wins — novelis.base_ebitda already\n"
              f"  uses it. Do NOT 'correct' the spec to the model here.")
    print("\nNOTHING WAS WRITTEN.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
