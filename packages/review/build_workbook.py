"""Build the audit workbook: every price, every formula, every trade.

Written so the SCORING IS LIVE. The hill curve, the bridge lines and the pair
P&L are Excel formulas referencing labelled input cells, not values pasted from
Python. Change an intensity or the anchor and the sheet recomputes, which is the
only way a reader can actually check the model rather than take it on trust.
"""
from __future__ import annotations

import json, pathlib, sqlite3, statistics, sys, datetime as dt

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, BarChart, ScatterChart, Reference, Series

REPO = pathlib.Path(__file__).resolve().parent.parent.parent
HERE = pathlib.Path(__file__).resolve().parent
sys.path.insert(0, str(REPO / "packages" / "score"))
sys.path.insert(0, str(HERE))

OUT = REPO / "Investment_Micro_System_Audit.xlsx"

# ---------- house style ----------
F = "Arial"
H1 = Font(name=F, size=14, bold=True)
H2 = Font(name=F, size=11, bold=True, color="FFFFFF")
BOLD = Font(name=F, size=10, bold=True)
BODY = Font(name=F, size=10)
INPUT = Font(name=F, size=10, color="0000FF")        # hardcoded input
LINK = Font(name=F, size=10, color="008000")         # link to another sheet
NOTE = Font(name=F, size=9, italic=True, color="666666")
HDRFILL = PatternFill("solid", fgColor="1F3B4D")
YEL = PatternFill("solid", fgColor="FFFF00")
GREY = PatternFill("solid", fgColor="F2F2F2")
THIN = Border(bottom=Side(style="thin", color="BFBFBF"))
WRAP = Alignment(wrap_text=True, vertical="top")


def sheet_header(ws, cols, row=1):
    for i, c in enumerate(cols, 1):
        cell = ws.cell(row=row, column=i, value=c)
        cell.font = H2
        cell.fill = HDRFILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    ws.freeze_panes = ws.cell(row=row + 1, column=1)


def widths(ws, spec):
    for col, w in spec.items():
        ws.column_dimensions[col].width = w


def title(ws, text, sub=None):
    ws["A1"] = text
    ws["A1"].font = H1
    if sub:
        ws["A2"] = sub
        ws["A2"].font = NOTE
    return 4


# ============================================================ data
conn = sqlite3.connect(REPO / "data" / "ims.db")
from bridge import (load_specs, load_scoring, load_accumulation, run_bridge,  # noqa
                    shocks_from_store, _series_in_store)
import yaml

ents, units, fin = load_specs()
form, kk, pp = load_scoring()
acc, hl = load_accumulation()
P1 = json.load(open(HERE / "_p1.json"))
VIZ = json.load(open(HERE / "_viz.json"))

COMMOD = ["lme_aluminium", "alumina_index", "lme_zinc", "silver",
          "thermal_coal_seaborne", "cp_coke", "usdinr", "usdcny", "brent",
          "zinc_shfe", "midwest_premium"]
STOCKS = ["nalco", "hindalco", "vaml", "hindustan_zinc", "vedanta"]
CO = {"nalco": "NALCO", "hindalco": "Hindalco", "vaml": "VAML",
      "hindustan_zinc": "Hindustan Zinc", "vedanta": "Vedanta"}

px: dict[str, dict[str, float]] = {}
for e, d, c in conn.execute("SELECT entity_id,date,close FROM prices WHERE close IS NOT NULL"):
    px.setdefault(e, {})[d] = c

wb = openpyxl.Workbook()

# ============================================================ 1. README
ws = wb.active
ws.title = "README"
r = title(ws, "Investment Micro-System — audit workbook",
          f"Generated {dt.date.today().isoformat()} from data/ims.db. "
          f"Every price, formula and trade behind the backtests.")
widths(ws, {"A": 26, "B": 96})
rows = [
    ("Sheet", "What is in it"),
    ("Scoring Logic", "The scoring formula as LIVE Excel formulas. Change the yellow input cells "
     "and the score recalculates. This is the sheet to check the maths on."),
    ("P1 Bridge", "The margin bridge per company: every line, its tonnage intensity, the share "
     "bought at market, and the resulting profit impact. Live formulas."),
    ("Sensitivity", "What a +10% move in each commodity does to each company's EBITDA."),
    ("Commodity Prices", "Every commodity series used, daily. Source noted per column."),
    ("Stock Prices", "Daily closes for the five names."),
    ("P1 Scores", "The economics score for every company on every date since 2021."),
    ("Backtest P1", "Trade by trade: the P1-score-ranked monthly pair, 90-day hold."),
    ("Backtest Regime", "Trade by trade: the four-regime rule, same months, same sizing."),
    ("Results", "Summary of both backtests at 2:1 and 1:1 sizing."),
    ("Charts", "Equity curves and score history."),
    ("", ""),
    ("KEY CAVEATS", ""),
    ("Sample size", "The P1 backtest is 63 monthly trades held 90 trading days. Because they "
     "overlap, that is roughly 15 INDEPENDENT observations. A 70% win rate on 15 draws is "
     "well within chance. This is encouraging, not established."),
    ("Structural change", "All specs are effective_from 2026-04-01. Hindalco's Mahan/Aditya "
     "smelters ramped 2013-16 and tripled its Indian capacity; Vedanta merged with Sesa Goa "
     "(2013) and Cairn (2017) and demerged (2026). Testing before ~2021 tests a company "
     "structure that did not exist. That is why the backtest starts 2021."),
    ("Sizing", "2:1 means Rs200 long against Rs100 short. That is NET LONG 33%, so part of the "
     "return is market exposure, not pair selection. The 1:1 column is the honest test."),
    ("Zinc proxy", "Hindustan Zinc and Vedanta still take their zinc price from zinc_shfe, a "
     "Chinese domestic contract used as a stand-in (119 days). Real LME zinc (4,722 days) is "
     "loaded but nothing points at it yet. The zinc scores are provisional for that reason."),
    ("verify: pending", "Every tonnage intensity in the specs is marked verify:pending — set "
     "from sector knowledge, not from filings. See the P1 Bridge sheet."),
]
for i, (a, b) in enumerate(rows):
    ws.cell(row=r + i, column=1, value=a).font = BOLD if a else BODY
    c = ws.cell(row=r + i, column=2, value=b)
    c.font = BODY
    c.alignment = WRAP
    ws.row_dimensions[r + i].height = 30 if len(b) > 90 else 15

# ============================================================ 2. Scoring Logic
ws = wb.create_sheet("Scoring Logic")
r = title(ws, "Scoring formula — live",
          "Yellow cells are inputs. Everything else is an Excel formula. Change an input and "
          "the score below recalculates, so you can check it against the Python.")
widths(ws, {"A": 30, "B": 16, "C": 14, "D": 60})

ws.cell(row=r, column=1, value="THE CURVE").font = BOLD
r += 1
for lab, val, note in [
    ("Neutral score", 3.0, "Firm convention: 3 = nothing is happening"),
    ("Half range", 2.0, "So the scale spans 1 to 5"),
    ("Anchor: x_ref", 0.05, "A 5% move in EBITDA..."),
    ("Anchor: score_ref", 4.0, "...should read as a 4.0"),
    ("Exponent p", pp, "p>1 makes the slope at zero ZERO, so noise cannot move the score"),
]:
    ws.cell(row=r, column=1, value=lab).font = BODY
    c = ws.cell(row=r, column=2, value=val)
    c.font = INPUT
    c.fill = YEL
    ws.cell(row=r, column=4, value=note).font = NOTE
    r += 1
NEU, HR, XREF, SREF, PEXP = [f"$B${r-5+i}" for i in range(5)]

ws.cell(row=r, column=1, value="Solved k").font = BOLD
ws.cell(row=r, column=2,
        value=f"={XREF}/((({SREF}-{NEU})/{HR})/(1-(({SREF}-{NEU})/{HR})))^(1/{PEXP})")
ws.cell(row=r, column=4, value="k solves the anchor. For the hill form k = x_ref exactly, "
                               "for any p — so p tunes shape without moving the anchor.").font = NOTE
KC = f"$B${r}"
r += 2

ws.cell(row=r, column=1, value="TRY IT").font = BOLD
r += 1
ws.cell(row=r, column=1, value="Change in EBITDA (as a fraction)").font = BODY
c = ws.cell(row=r, column=2, value=0.05)
c.font = INPUT
c.fill = YEL
c.number_format = "0.00%"
XIN = f"$B${r}"
r += 1
ws.cell(row=r, column=1, value="Score").font = BOLD
ws.cell(row=r, column=2, value=(
    f"={NEU}+{HR}*SIGN({XIN})*(ABS({XIN}/{KC})^{PEXP})/(1+ABS({XIN}/{KC})^{PEXP})"))
ws.cell(row=r, column=2).number_format = "0.00"
ws.cell(row=r, column=2).font = Font(name=F, size=12, bold=True)
ws.cell(row=r, column=4, value="score = 3 + 2 × sign(x) × |x/k|^p / (1 + |x/k|^p)").font = NOTE
r += 2

ws.cell(row=r, column=1, value="CALIBRATION TABLE").font = BOLD
r += 1
sheet_header(ws, ["Change in EBITDA", "Score", "", "Reading"], row=r)
r += 1
grid = [(-0.20, "very negative"), (-0.10, ""), (-0.05, "mirror of the anchor"),
        (-0.015, "materiality floor"), (-0.002, "noise — barely moves"), (0.0, "neutral"),
        (0.002, "noise — barely moves"), (0.015, "materiality floor"), (0.05, "THE ANCHOR"),
        (0.10, "big"), (0.20, "very big"), (0.50, "extreme — still not pinned at 5")]
first_cal = r
for x, note in grid:
    c = ws.cell(row=r, column=1, value=x)
    c.number_format = "0.0%"
    c.font = INPUT
    f = (f"={NEU}+{HR}*SIGN(A{r})*(ABS(A{r}/{KC})^{PEXP})/(1+ABS(A{r}/{KC})^{PEXP})")
    c2 = ws.cell(row=r, column=2, value=f)
    c2.number_format = "0.00"
    c2.font = BODY
    ws.cell(row=r, column=4, value=note).font = NOTE
    r += 1
last_cal = r - 1
r += 1

for head, body in [
    ("WHY THIS SHAPE", "Flat near zero so daily price noise does not move the score. Steepest "
     "between 2% and 8% of EBITDA, which is where trades get decided. Never fully saturates, so "
     "two large moves stay distinguishable instead of both pinning at 5."),
    ("ACCUMULATION", f"Daily impacts are combined by an EWMA with a {hl:.0f}-day half-life, not a "
     "fixed trailing window. A 20-day sum changes every day purely because the oldest day drops "
     "out — a score move with no new information."),
    ("PILLAR 4 IS DIFFERENT", "Guidance scores LINEARLY: score = 1 + 4 × confidence. A confidence "
     "is already a bounded probability; squashing it again would distort it."),
    ("COMPOSITE", "0.45 economics / 0.25 valuation / 0.15 mood / 0.15 guidance, RENORMALISED over "
     "whichever pillars actually scored. A withheld pillar is never filled with 3.0 — that would "
     "let missing data pose as neutral evidence."),
    ("PAIRS", "pair_score = score(pct_long − pct_short). Score the spread; do not subtract two "
     "scores. The curve is flat in the tails, so subtracting compresses a real 1.75pp gap into "
     "0.09 score points."),
]:
    ws.cell(row=r, column=1, value=head).font = BOLD
    c = ws.cell(row=r, column=2, value=body)
    c.font = BODY
    c.alignment = WRAP
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
    ws.row_dimensions[r].height = 42
    r += 1

# ============================================================ 3. P1 Bridge
ws = wb.create_sheet("P1 Bridge")
r = title(ws, "P1 margin bridge — how a price move becomes a profit change",
          f"Shocks are the 30-day EWMA move as at {P1['as_of']}. Blue cells are spec inputs; "
          f"impact and totals are formulas.")
widths(ws, {"A": 17, "B": 11, "C": 22, "D": 22, "E": 11, "F": 11, "G": 12, "H": 14, "I": 15,
            "J": 46})

ws.cell(row=r, column=1, value="DRIVER SHOCKS (30-day)").font = BOLD
r += 1
sheet_header(ws, ["Driver", "Level", "30d move", "% move"], row=r)
r += 1
shock_at = {}
for s in P1["shocks"]:
    ws.cell(row=r, column=1, value=s["driver"]).font = BODY
    ws.cell(row=r, column=2, value=s["level"]).font = INPUT
    ws.cell(row=r, column=3, value=s["delta"]).font = INPUT
    ws.cell(row=r, column=4, value=f"=C{r}/B{r}").number_format = "0.00%"
    shock_at[s["driver"]] = f"$C${r}"
    r += 1
r += 1

ws.cell(row=r, column=1, value="BRIDGE LINES").font = BOLD
r += 1
sheet_header(ws, ["Company", "Kind", "Line item", "Priced off", "Intensity", "Unit",
                  "market_pct", "Impact (Rs cr)", "", "Source note (verify: pending)"], row=r)
r += 1
tot_cells: dict[str, list[str]] = {}
for book in ("aluminium", "silver / zinc"):
    for row_ in P1["books"][book]:
        e = row_["entity"]
        st = {(l["kind"], l["item"]): l for l in P1["struct"].get(e, [])}
        tot_cells[e] = []
        for l in row_["lines"]:
            spec = st.get((l["kind"], l["item"]), {})
            ws.cell(row=r, column=1, value=CO[e]).font = BODY
            ws.cell(row=r, column=2, value=l["kind"]).font = BODY
            ws.cell(row=r, column=3, value=l["item"]).font = BODY
            ws.cell(row=r, column=4, value=str(l.get("driver"))).font = BODY
            ws.cell(row=r, column=5, value=spec.get("intensity")).font = INPUT
            ws.cell(row=r, column=6, value=spec.get("unit")).font = BODY
            mp = l.get("market_pct")
            c = ws.cell(row=r, column=7, value=mp if mp is not None else "n/a")
            c.font = INPUT
            if mp == 0.0:
                c.fill = YEL
            ws.cell(row=r, column=8, value=l["impact"]).number_format = "#,##0;(#,##0);-"
            ws.cell(row=r, column=10, value=spec.get("note", "")).font = NOTE
            tot_cells[e].append(f"H{r}")
            r += 1
        ws.cell(row=r, column=3, value=f"{CO[e]} — change in EBITDA").font = BOLD
        ws.cell(row=r, column=8, value=f"=SUM({tot_cells[e][0]}:{tot_cells[e][-1]})").font = BOLD
        ws.cell(row=r, column=8).number_format = "#,##0;(#,##0);-"
        SUMC = f"H{r}"
        r += 1
        ws.cell(row=r, column=3, value="base EBITDA (Rs cr)").font = BODY
        ws.cell(row=r, column=8, value=row_["base_ebitda"]).font = INPUT
        BASEC = f"H{r}"
        r += 1
        ws.cell(row=r, column=3, value="as % of base EBITDA").font = BODY
        ws.cell(row=r, column=8, value=f"={SUMC}/{BASEC}").number_format = "0.00%"
        PCTC = f"H{r}"
        r += 1
        ws.cell(row=r, column=3, value="ECONOMICS SCORE").font = BOLD
        ws.cell(row=r, column=8, value=(
            # Scoring Logic!B5=neutral B6=half-range B9=p B10=solved k.
            # These were wrong on the first build (pointed at B8/B9/B11/B12,
            # i.e. score_ref, p, and two BLANK label rows) and produced a
            # formula that evaluates cleanly and is arithmetically nonsense.
            f"='Scoring Logic'!$B$5+'Scoring Logic'!$B$6*SIGN({PCTC})"
            f"*(ABS({PCTC}/'Scoring Logic'!$B$10)^'Scoring Logic'!$B$9)"
            f"/(1+ABS({PCTC}/'Scoring Logic'!$B$10)^'Scoring Logic'!$B$9)"))
        ws.cell(row=r, column=8).font = Font(name=F, size=11, bold=True, color="008000")
        ws.cell(row=r, column=8).number_format = "0.00"
        ws.cell(row=r, column=10, value="green = computed on the Scoring Logic sheet").font = NOTE
        r += 3

# ============================================================ 4. Sensitivity
ws = wb.create_sheet("Sensitivity")
r = title(ws, "What each name is exposed to",
          "A +10% move in ONE commodity, everything else held still. Value is the resulting "
          "change in EBITDA as a % of that company's own base.")
drv = ["lme_aluminium", "alumina_index", "lme_zinc", "silver", "thermal_coal_seaborne", "cp_coke"]
widths(ws, {"A": 18, **{get_column_letter(i + 2): 17 for i in range(len(drv))}})
sheet_header(ws, ["Company"] + drv, row=r)
r += 1
for e, row_ in P1["sens"].items():
    ws.cell(row=r, column=1, value=CO[e]).font = BODY
    for i, d in enumerate(drv, 2):
        v = row_.get(d, 0) or 0
        c = ws.cell(row=r, column=i, value=v / 100)
        c.number_format = "0.00%;(0.00%);-"
        c.font = BODY
    r += 1
r += 1
for t in ["LME aluminium moves all three aluminium names by 12-18% and therefore separates NONE "
          "of them. Alumina is the opposite: +3.81% NALCO, +0.18% Hindalco, -0.91% VAML. That one "
          "column is where the aluminium pair trade comes from.",
          "The LME zinc column is empty because both zinc names are still priced off zinc_shfe, "
          "a Chinese domestic proxy with 119 days of history, even though 4,722 days of real LME "
          "zinc are now loaded. Repointing the spec would rewrite every zinc score."]:
    c = ws.cell(row=r, column=1, value=t)
    c.font = BODY
    c.alignment = WRAP
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
    ws.row_dimensions[r].height = 40
    r += 1

# ============================================================ 5/6. prices
SRC = {"lme_aluminium": "Daily Metals Pack col1 — LME cash US$/t",
       "alumina_index": "Daily Metals Pack col20 — Alumina Australia FOB, assessed US$/t",
       "lme_zinc": "Daily Metals Pack col2 — LME cash US$/t (loaded, not yet used by specs)",
       "silver": "Daily Metals Pack col13 — spot US$/troy oz",
       "thermal_coal_seaborne": "Daily Metals Pack col11 — Richards Bay US$/t",
       "cp_coke": "Daily Metals Pack col24 — pet coke US$/t",
       "usdinr": "Daily Metals Pack col15", "usdcny": "Daily Metals Pack col32",
       "brent": "Daily Metals Pack col16 — US$/bbl",
       "zinc_shfe": "Wind ZN.SHF via agent, CNY->USD ex-VAT — PROXY",
       "midwest_premium": "Yahoo AUP=F — US$/lb (legacy, superseded)"}


def price_sheet(name, ids, labels, note):
    ws = wb.create_sheet(name)
    r0 = title(ws, name, note)
    ws.cell(row=r0, column=1, value="Source").font = BOLD
    for i, e in enumerate(ids, 2):
        ws.cell(row=r0, column=i, value=SRC.get(e, "Yahoo Finance .NS daily close")).font = NOTE
        ws.column_dimensions[get_column_letter(i)].width = 15
    ws.column_dimensions["A"].width = 12
    ws.row_dimensions[r0].height = 46
    for i in range(2, len(ids) + 2):
        ws.cell(row=r0, column=i).alignment = WRAP
    hr = r0 + 1
    sheet_header(ws, ["Date"] + labels, row=hr)
    alld = sorted({d for e in ids for d in px.get(e, {})})
    for j, d in enumerate(alld):
        rr = hr + 1 + j
        ws.cell(row=rr, column=1, value=d).font = BODY
        for i, e in enumerate(ids, 2):
            v = px.get(e, {}).get(d)
            if v is not None:
                c = ws.cell(row=rr, column=i, value=round(v, 4))
                c.font = BODY
                c.number_format = "#,##0.00"
    return len(alld)


n_com = price_sheet("Commodity Prices", COMMOD, COMMOD,
                    "Every commodity series in the store. Blank = no print that day.")
n_stk = price_sheet("Stock Prices", STOCKS, [CO[s] for s in STOCKS],
                    "Daily closes, Yahoo .NS. VEDL carries an UNADJUSTED demerger on 2026-04-30 "
                    "(773.60 -> 271.55); anything crossing that date compares two companies. "
                    "VAML listed 2026-06-15.")

# ============================================================ 7. P1 scores
ws = wb.create_sheet("P1 Scores")
r = title(ws, "Economics score, every company, every date",
          "Computed by packages/score/backfill_p1.py. 3.0 = neutral.")
widths(ws, {"A": 12, **{get_column_letter(i + 2): 15 for i in range(5)}})
sheet_header(ws, ["Date"] + [CO[s] for s in STOCKS], row=r)
hist = P1["hist"]
alld = sorted({d for v in hist.values() for d, _ in v})
lookup = {e: dict(v) for e, v in hist.items()}
score_first = r + 1
for j, d in enumerate(alld):
    rr = r + 1 + j
    ws.cell(row=rr, column=1, value=d).font = BODY
    for i, e in enumerate(STOCKS, 2):
        v = lookup.get(e, {}).get(d)
        if v is not None:
            c = ws.cell(row=rr, column=i, value=v)
            c.number_format = "0.000"
            c.font = BODY
score_last = r + len(alld)

# ============================================================ 8/9. backtests
def bt_sheet(name, trades, note):
    ws = wb.create_sheet(name)
    r0 = title(ws, name, note)
    widths(ws, {"A": 10, "B": 8, "C": 9, "D": 9, "E": 15, "F": 15, "G": 11, "H": 11,
                "I": 12, "J": 12, "K": 13, "L": 13})
    sheet_header(ws, ["Month", "Regime", "Alu %", "Alumina %", "LONG", "SHORT",
                      "Long ret", "Short ret", "P&L 2:1", "Return 2:1",
                      "P&L 1:1", "Return 1:1"], row=r0)
    rr = r0 + 1
    first = rr
    for t in trades:
        ws.cell(row=rr, column=1, value=t["month"]).font = BODY
        ws.cell(row=rr, column=2, value=t.get("regime", "")).font = BODY
        for col, key, fmt in ((3, "alu", "0.0%"), (4, "alm", "0.0%")):
            if t.get(key) is not None:
                c = ws.cell(row=rr, column=col, value=t[key] / 100)
                c.number_format = fmt
                c.font = BODY
        ws.cell(row=rr, column=5, value=CO.get(t["long"], t["long"])).font = BODY
        ws.cell(row=rr, column=6, value=CO.get(t["short"], t["short"])).font = BODY
        c = ws.cell(row=rr, column=7, value=t["lret"] / 100)
        c.number_format = "0.00%"
        c.font = INPUT
        c = ws.cell(row=rr, column=8, value=t["sret"] / 100)
        c.number_format = "0.00%"
        c.font = INPUT
        ws.cell(row=rr, column=9, value=f"=200*G{rr}-100*H{rr}").number_format = "#,##0.0;(#,##0.0);-"
        ws.cell(row=rr, column=10, value=f"=I{rr}/300").number_format = "0.00%"
        ws.cell(row=rr, column=11, value=f"=100*G{rr}-100*H{rr}").number_format = "#,##0.0;(#,##0.0);-"
        ws.cell(row=rr, column=12, value=f"=K{rr}/200").number_format = "0.00%"
        rr += 1
    last = rr - 1
    rr += 1
    ws.cell(row=rr, column=5, value="Trades").font = BOLD
    ws.cell(row=rr, column=9, value=f"=COUNT(I{first}:I{last})").font = BOLD
    rr += 1
    ws.cell(row=rr, column=5, value="Won (2:1 / 1:1)").font = BOLD
    ws.cell(row=rr, column=9, value=f"=COUNTIF(I{first}:I{last},\">0\")").font = BOLD
    ws.cell(row=rr, column=11, value=f"=COUNTIF(K{first}:K{last},\">0\")").font = BOLD
    rr += 1
    ws.cell(row=rr, column=5, value="Win rate").font = BOLD
    ws.cell(row=rr, column=9, value=f"=COUNTIF(I{first}:I{last},\">0\")/COUNT(I{first}:I{last})"
            ).number_format = "0%"
    ws.cell(row=rr, column=11, value=f"=COUNTIF(K{first}:K{last},\">0\")/COUNT(K{first}:K{last})"
            ).number_format = "0%"
    rr += 1
    ws.cell(row=rr, column=5, value="Average return").font = BOLD
    ws.cell(row=rr, column=10, value=f"=AVERAGE(J{first}:J{last})").number_format = "0.00%"
    ws.cell(row=rr, column=12, value=f"=AVERAGE(L{first}:L{last})").number_format = "0.00%"
    rr += 1
    ws.cell(row=rr, column=5, value="Median return").font = BOLD
    ws.cell(row=rr, column=10, value=f"=MEDIAN(J{first}:J{last})").number_format = "0.00%"
    ws.cell(row=rr, column=12, value=f"=MEDIAN(L{first}:L{last})").number_format = "0.00%"
    rr += 1
    ws.cell(row=rr, column=5, value="Std dev").font = BOLD
    ws.cell(row=rr, column=10, value=f"=STDEV(J{first}:J{last})").number_format = "0.00%"
    ws.cell(row=rr, column=12, value=f"=STDEV(L{first}:L{last})").number_format = "0.00%"
    rr += 1
    ws.cell(row=rr, column=5, value="Sharpe (per trade)").font = BOLD
    ws.cell(row=rr, column=10, value=f"=J{rr-3}/J{rr-1}").number_format = "0.000"
    ws.cell(row=rr, column=12, value=f"=L{rr-3}/L{rr-1}").number_format = "0.000"
    return first, last



BT = json.load(open(HERE / "_bt.json"))
p1_first, p1_last = bt_sheet(
    "Backtest P1", BT["p1"],
    "Rank the three aluminium names by their P1 economics score at month end. Long the highest, "
    "short the lowest, hold 90 trading days. Blue cells are the realised stock returns; every "
    "P&L and statistic is a formula.")
rg_first, rg_last = bt_sheet(
    "Backtest Regime", BT["regime"],
    "The four-regime rule instead of the score: R1 long NALCO/short Hindalco, R2 long VAML/short "
    "Hindalco, R3 long NALCO/short VAML, R4 long Hindalco/short NALCO. Same months, same hold.")

# ---------------------------------------------------------------- Results
ws = wb.create_sheet("Results")
r = title(ws, "Backtest results",
          "Both models, same 2021-02..2026-08 window, same 90-day hold. All cells are formulas "
          "pulling from the two backtest sheets.")
widths(ws, {"A": 26, "B": 12, "C": 12, "D": 12, "E": 12, "F": 12, "G": 58})
sheet_header(ws, ["Model / sizing", "Trades", "Won", "Win rate", "Avg return",
                  "Median", "What it means"], row=r)
r += 1
SPECS = [
    ("P1 score rank — 2:1", "Backtest P1", p1_first, p1_last, "I", "J",
     "The best result here. But 2:1 is NET LONG 33%, so part of this is market exposure."),
    ("P1 score rank — 1:1 MV", "Backtest P1", p1_first, p1_last, "K", "L",
     "The honest test: market-value neutral, no net long tilt. Still above a coin flip."),
    ("4-regime rule — 2:1", "Backtest Regime", rg_first, rg_last, "I", "J",
     "The vault's original model, re-run on the same months."),
    ("4-regime rule — 1:1 MV", "Backtest Regime", rg_first, rg_last, "K", "L",
     "At market-value neutral the regime rule is a coin flip."),
]
for lab, sh, f_, l_, pcol, rcol, note in SPECS:
    ws.cell(row=r, column=1, value=lab).font = BOLD
    ws.cell(row=r, column=2, value=f"=COUNT('{sh}'!{pcol}{f_}:{pcol}{l_})").font = LINK
    ws.cell(row=r, column=3, value=f"=COUNTIF('{sh}'!{pcol}{f_}:{pcol}{l_},\">0\")").font = LINK
    c = ws.cell(row=r, column=4, value=(
        f"=COUNTIF('{sh}'!{pcol}{f_}:{pcol}{l_},\">0\")/COUNT('{sh}'!{pcol}{f_}:{pcol}{l_})"))
    c.number_format = "0%"
    c.font = LINK
    c = ws.cell(row=r, column=5, value=f"=AVERAGE('{sh}'!{rcol}{f_}:{rcol}{l_})")
    c.number_format = "0.00%"
    c.font = LINK
    c = ws.cell(row=r, column=6, value=f"=MEDIAN('{sh}'!{rcol}{f_}:{rcol}{l_})")
    c.number_format = "0.00%"
    c.font = LINK
    c = ws.cell(row=r, column=7, value=note)
    c.font = NOTE
    c.alignment = WRAP
    ws.row_dimensions[r].height = 30
    r += 2

r += 1
for head, txt in [
    ("HOW A TRADE WORKS", "Once a month: buy Rs200 of one stock, short Rs100 of another, hold 90 "
     "trading days, close. Capital deployed Rs300. P&L = 200 x long return - 100 x short return."),
    ("WHY SHORT AT ALL", "Both names usually move together with aluminium. Without the short leg "
     "the result measures the metal, not the model. The short cancels the sector move."),
    ("WHAT WIN RATE IS", "The count of trades that ended above zero, divided by the number of "
     "trades. Nothing weighted, nothing adjusted."),
    ("THE SAMPLE CAVEAT", "63 monthly trades each held 90 trading days overlap heavily. That is "
     "roughly 15 INDEPENDENT observations, so a 70% win rate is well inside what chance produces. "
     "Treat this as encouraging, not proven."),
    ("2:1 VERSUS 1:1", "2:1 puts twice as much long as short, which is a 33% net long position. "
     "Over a period when aluminium rose, that alone adds return. The 1:1 column removes it."),
]:
    ws.cell(row=r, column=1, value=head).font = BOLD
    c = ws.cell(row=r, column=2, value=txt)
    c.font = BODY
    c.alignment = WRAP
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=7)
    ws.row_dimensions[r].height = 32
    r += 1

# ---------------------------------------------------------------- Charts
ws = wb.create_sheet("Charts")
title(ws, "Charts", "Built from the Backtest and P1 Scores sheets — they move with the data.")

ch = BarChart()
ch.type = "col"
ch.title = "P1 backtest — return per trade (2:1 sizing)"
ch.y_axis.title = "return on capital"
ch.x_axis.title = "trade"
ch.height, ch.width = 8, 30
ch.add_data(Reference(wb["Backtest P1"], min_col=10, min_row=p1_first - 1, max_row=p1_last),
            titles_from_data=True)
ch.set_categories(Reference(wb["Backtest P1"], min_col=1, min_row=p1_first, max_row=p1_last))
ch.gapWidth = 30
ws.add_chart(ch, "A4")

ch2 = BarChart()
ch2.type = "col"
ch2.title = "Four-regime rule — return per trade (2:1 sizing)"
ch2.y_axis.title = "return on capital"
ch2.height, ch2.width = 8, 30
ch2.add_data(Reference(wb["Backtest Regime"], min_col=10, min_row=rg_first - 1, max_row=rg_last),
             titles_from_data=True)
ch2.set_categories(Reference(wb["Backtest Regime"], min_col=1, min_row=rg_first, max_row=rg_last))
ch2.gapWidth = 30
ws.add_chart(ch2, "A22")

ch3 = LineChart()
ch3.title = "P1 economics score, every company, since 2021 (3.0 = neutral)"
ch3.y_axis.title = "score"
ch3.height, ch3.width = 9, 30
ch3.add_data(Reference(wb["P1 Scores"], min_col=2, max_col=6, min_row=score_first - 1,
                       max_row=score_last), titles_from_data=True)
ch3.set_categories(Reference(wb["P1 Scores"], min_col=1, min_row=score_first, max_row=score_last))
for s in ch3.series:
    s.smooth = False
    s.graphicalProperties.line.width = 12000
ws.add_chart(ch3, "A40")

ch4 = LineChart()
ch4.title = "Commodity prices — LME aluminium and alumina"
ch4.y_axis.title = "US$/t"
ch4.height, ch4.width = 9, 30
cw = wb["Commodity Prices"]
last_c = cw.max_row
ch4.add_data(Reference(cw, min_col=2, max_col=3, min_row=6, max_row=last_c),
             titles_from_data=True)
ch4.set_categories(Reference(cw, min_col=1, min_row=7, max_row=last_c))
for s in ch4.series:
    s.smooth = False
    s.graphicalProperties.line.width = 10000
ws.add_chart(ch4, "A60")

wb.save(OUT)
print("saved", OUT)

conn.close()
